from openpyxl import load_workbook
import pyperclip

file_location = './subdepartment_list.xlsx'

workbook = load_workbook(filename = file_location)
sheet = workbook.active

departments = []
mbu = int(input('mbu: '))

for row in sheet.iter_rows(min_row=4, min_col=5, max_col=7,values_only=True):
    if row[0] == mbu:
        departments.append(row[2])
        
result = ', '.join(map(str, set(departments)))
pyperclip.copy(result)

print(f'\nThe following has been clipped to the system clipboard:\n{result}')
